import io
from datetime import datetime
from urllib.parse import quote

from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from database import get_db

router = APIRouter()


HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(fill_type="solid", start_color="1D1B31", end_color="1D1B31")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def style_header(ws):
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER


def auto_adjust_columns(ws):
    for column_cells in ws.columns:
        width = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, len(value))
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(width + 4, 12), 40)


def sanitize_sheet_name(name: str, existing_names: set[str]) -> str:
    cleaned = "".join(ch for ch in (name or "未命名课程") if ch not in '\\/?*[]:')
    cleaned = cleaned.strip() or "未命名课程"
    cleaned = cleaned[:31]
    base = cleaned
    counter = 1
    while cleaned in existing_names:
        suffix = f"_{counter}"
        cleaned = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    existing_names.add(cleaned)
    return cleaned


@router.get("/export/excel")
def export_excel():
    conn = get_db()
    cursor = conn.cursor()

    workbook = Workbook()
    workbook.remove(workbook.active)
    sheet_names = set()

    cursor.execute("SELECT COUNT(1) AS cnt FROM student")
    student_count = int(cursor.fetchone()["cnt"])
    cursor.execute("SELECT COUNT(1) AS cnt FROM course")
    course_count = int(cursor.fetchone()["cnt"])
    cursor.execute("SELECT COUNT(1) AS cnt FROM attendance")
    attendance_count = int(cursor.fetchone()["cnt"])
    cursor.execute("SELECT COUNT(1) AS cnt FROM group_list WHERE is_active = 1")
    group_count = int(cursor.fetchone()["cnt"])
    cursor.execute("SELECT COUNT(1) AS cnt FROM enrollment")
    enrollment_count = int(cursor.fetchone()["cnt"])
    cursor.execute("SELECT COUNT(DISTINCT attend_date) AS cnt FROM attendance")
    attendance_day_count = int(cursor.fetchone()["cnt"])
    cursor.execute("SELECT MIN(attend_date) AS min_date, MAX(attend_date) AS max_date FROM attendance")
    attendance_span = cursor.fetchone()

    exported_at = datetime.now()

    ws_meta = workbook.create_sheet("元数据")
    sheet_names.add("元数据")
    ws_meta.append(["指标", "数据", "说明"])
    style_header(ws_meta)
    meta_rows = [
        ("系统名称", "爱乐之城数据导出", "用于归档学生、课程、团和签到信息"),
        ("导出时间", exported_at.strftime("%Y-%m-%d %H:%M:%S"), "本次导出的服务器时间"),
        ("文件名", f"爱乐之城数据_{exported_at.strftime('%Y%m%d-%H%M%S')}.xlsx", "Excel 文件命名规则"),
        ("学生总数", student_count, "当前系统内学生数量"),
        ("课程总数", course_count, "当前系统内课程数量"),
        ("团总数", group_count, "当前启用的团数量"),
        ("报名总数", enrollment_count, "所有课程报名关系总数"),
        ("签到总次数", attendance_count, "所有签到记录累计条数"),
        ("签到日期总数", attendance_day_count, "去重后的签到日期总数"),
        (
            "签到时间范围",
            f"{attendance_span['min_date'] or '无'} ~ {attendance_span['max_date'] or '无'}",
            "系统中最早到最晚的签到日期",
        ),
    ]
    for row in meta_rows:
        ws_meta.append(list(row))
    for row in ws_meta.iter_rows(min_row=2):
        row[0].alignment = LEFT
        row[1].alignment = LEFT
        row[2].alignment = LEFT
    auto_adjust_columns(ws_meta)

    ws_students = workbook.create_sheet("学生列表")
    sheet_names.add("学生列表")
    ws_students.append(["学号", "姓名", "团"])
    style_header(ws_students)
    cursor.execute(
        """
        SELECT student_no, name, COALESCE(group_name, '') AS group_name
        FROM student
        ORDER BY COALESCE(group_name, ''), student_no
        """
    )
    for row in cursor.fetchall():
        ws_students.append([row["student_no"], row["name"], row["group_name"] or "未分组"])
    auto_adjust_columns(ws_students)

    ws_courses = workbook.create_sheet("课程列表")
    sheet_names.add("课程列表")
    ws_courses.append(["课程名", "老师", "开课时间", "总课时"])
    style_header(ws_courses)
    cursor.execute(
        """
        SELECT name, COALESCE(teacher, '') AS teacher, start_date, total_lessons
        FROM course
        ORDER BY id
        """
    )
    courses = [dict(row) for row in cursor.fetchall()]
    for row in courses:
        ws_courses.append([
            row["name"],
            row["teacher"] or "未设置",
            row["start_date"] or "未设置",
            row["total_lessons"],
        ])
    auto_adjust_columns(ws_courses)

    ws_groups = workbook.create_sheet("团列表")
    sheet_names.add("团列表")
    ws_groups.append(["团名称"])
    style_header(ws_groups)
    cursor.execute("SELECT name FROM group_list WHERE is_active = 1 ORDER BY sort_order ASC, name ASC")
    for row in cursor.fetchall():
        ws_groups.append([row["name"]])
    auto_adjust_columns(ws_groups)

    cursor.execute("SELECT id, name FROM course ORDER BY id")
    all_courses = [dict(row) for row in cursor.fetchall()]
    for course in all_courses:
        ws_course = workbook.create_sheet(sanitize_sheet_name(course["name"], sheet_names))

        cursor.execute(
            """
            SELECT s.id, s.student_no, s.name
            FROM enrollment e
            JOIN student s ON s.id = e.student_id
            WHERE e.course_id = ?
            ORDER BY s.student_no
            """,
            (course["id"],),
        )
        students = [dict(row) for row in cursor.fetchall()]

        cursor.execute(
            """
            SELECT DISTINCT attend_date
            FROM attendance
            WHERE course_id = ?
            ORDER BY attend_date
            """,
            (course["id"],),
        )
        dates = [row["attend_date"] for row in cursor.fetchall()]

        headers = ["学生学号", "姓名"] + dates
        ws_course.append(headers)
        style_header(ws_course)

        attendance_lookup = set()
        if dates:
            cursor.execute(
                """
                SELECT student_id, attend_date
                FROM attendance
                WHERE course_id = ?
                """,
                (course["id"],),
            )
            attendance_lookup = {(row["student_id"], row["attend_date"]) for row in cursor.fetchall()}

        for student in students:
            row = [student["student_no"], student["name"]]
            for attend_date in dates:
                row.append("√" if (student["id"], attend_date) in attendance_lookup else "")
            ws_course.append(row)

        for row in ws_course.iter_rows(min_row=2):
            for idx, cell in enumerate(row):
                cell.alignment = LEFT if idx < 2 else CENTER
        auto_adjust_columns(ws_course)

    conn.close()

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"爱乐之城数据_{exported_at.strftime('%Y%m%d-%H%M%S')}.xlsx"
    encoded_filename = quote(filename)
    headers = {
        "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
        "Access-Control-Expose-Headers": "Content-Disposition",
    }
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
